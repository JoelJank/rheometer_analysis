import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import savgol_filter
import math
import os 
import sys
import pandas as pd
from openpyxl import load_workbook
import json
settings_path = "settings/settings.json"
excel_path = "data/20250707Glass12_witharray.xlsx"

def get_sheetnames(Filepath):
    wb = load_workbook(Filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def get_settings_file(Settings_path):
    with open(Settings_path,'r') as f:
        return json.load(f)

def get_excel_file(Excel_path, Sheet_name):
    file = pd.read_excel(Excel_path, sheet_name = Sheet_name,header=0, skiprows=[1,2])
    mask = file.map(lambda x: isinstance(x, str)).any(axis=1)
    file = file[~mask].reset_index(drop=True)
    return file

def lin_fit(m, n, x):
    return m * x + n


settings = get_settings_file(settings_path)
plt.style.use(settings["plotstyle"])

sheetnames = get_sheetnames(settings["input_file"])
del sheetnames[:settings["skip_sheets"]]
del sheetnames[:-settings["ignore_sheets_end"]]


processed_data = []

for sheet in sheetnames:
    current_data = []
    df = get_excel_file(settings["input_file"], sheet)
    time_col = df["Time"].values #Determine switch between different consolidation forces
    jumps_where = np.where(np.diff(time_col) > 10)[0]
    jumps_end = jumps_where[-1]+jumps_where[0]+1
    jumps = np.append(jumps_where, jumps_end)
    jump_starts = np.array([0])
    for i in range(len(jumps)-1):
        jump_starts = np.append(jump_starts, jumps[i]+1)


    fig, ax = plt.subplots(figsize=(12,5))
    ax2 = ax.twinx()
    ax3 = ax.twinx()
    ax3.spines.right.set_position(("axes",1.1))
    
    
    
    #Normal Stress with mean
    p1, = ax.plot(df["Time"],df["Normal Stress"],marker="o",linestyle="None",markersize=1, color="blue",label="Normal Stress")
    
    for i in range(len(jump_starts)):
        mean = np.mean(df["Normal Stress"][jump_starts[i]:jumps[i]])
        ax.hlines(mean, xmin = df["Time"][jump_starts[i]], xmax = df["Time"][jumps[i]], color="red", linestyle="--")
        current_data.append(mean)
    ax.set_ylabel("Normal Stress [Pa]"); ax.set_xlabel("Time [s]")
    ax.yaxis.label.set_color(p1.get_color())
    ax.tick_params(axis = 'y', colors = p1.get_color())
    ax.set_title(f"{sheet}")
    lines1, labels1 = ax.get_legend_handles_labels()
    
    
    
    #Shear Stress with savgol filter
    p2, = ax2.plot(df["Time"],df["Shear Stress"],marker="o",linestyle = "None", markersize=1, color="orange", label="Shear Stress")
    for i in range(len(jump_starts)):
        sav = savgol_filter(df["Shear Stress"][jump_starts[i]:jumps[i]], settings["savgol_window_length"], settings["savgol_polyorder"])
        p2_sav, = ax2.plot(df["Time"][jump_starts[i]:jumps[i]], sav, color="purple", linestyle = "--")
        sav_max = [np.max(sav), list(sav).index(max(sav))]
        p2_max, = ax2.plot(df["Time"][jump_starts[i]+sav_max[1]], sav[sav_max[1]], marker="x", color = "red", linestyle = "None")
        current_data.append(sav[sav_max[1]])
    ax2.set_ylabel("Shear Stress [Pa]")
    ax2.yaxis.label.set_color(p2.get_color())
    ax2.tick_params(axis = 'y', colors = p2.get_color())
    ax2.grid()
    lines2, labels2 = ax2.get_legend_handles_labels()
    
    
    #Gap
    p3, = ax3.plot(df["Time"], df["Gap"], marker = "o", linestyle = "None", markersize = 1, color = "green", label = "Gap[mm]")
    ax3.set_ylabel("Gap [mm]")
    ax3.yaxis.label.set_color(p3.get_color())
    ax3.tick_params(axis = 'y', colors = p3.get_color())
    lines3, labels3 = ax3.get_legend_handles_labels()
    
    #Legend
    ax2.legend(lines1 + lines2 + lines3, labels1+labels2+labels3, loc='upper center', bbox_to_anchor=(0.5, 1.2), ncol=3, markerscale = 6)
    
    plt.savefig(f"{settings['output_folder']}/{sheet}_original_data.png", dpi=300, bbox_inches='tight')
    
    #Daten für Mohr-Coulomb speichern
    current_data = np.array(current_data)
    current_data = np.reshape(current_data, (2, 3))
    processed_data.append(current_data)
print(len(processed_data))
    
    
#Mohr-Coulomb
fit_data= []
plt.figure(figsize=(12,5))
    
for i in range(len(processed_data)):
    plot = plt.plot(processed_data[i][0],processed_data[i][1], linestyle = "None", marker = "x", label = f"{sheetnames[i]}")
    m, n = np.polyfit(processed_data[i][0], processed_data[i][1], 1)
    fit_data.append([m, n])
    x = np.linspace(0, max(processed_data[i][0]), 100)
    plt.plot(x, lin_fit(m, n, x), linestyle = "--", color = plot[0].get_color())
plt.xlabel("Normal Shear [Pa]"); plt.ylabel("Shear Stress [Pa]")
plt.title(f"{settings['title_of_experiment']} Mohr-Coulomb")
plt.legend(loc='upper left')
plt.grid()
plt.savefig(f"{settings['output_folder']}/{settings['title_of_experiment']}_Mohr-Coulomb.png", dpi=300, bbox_inches='tight')



    
    
    





